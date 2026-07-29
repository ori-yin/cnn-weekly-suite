"""
data.py - CNN Weekly Suite：数据读取与清洗（两模式共用）
参考 mcd-content-rank 的数据读取方式
"""

import logging
import numpy as np
import pandas as pd
from io import BytesIO
from shared.theme import COLUMN_MAPPING, NUMERIC_COLS, ENCODINGS

logger = logging.getLogger(__name__)


def _fuzzy_match_columns(df: pd.DataFrame) -> dict:
    """
    Fuzzy match 源文件列名到标准字段名。
    返回 {标准字段名: 源列名} 的映射字典。
    """
    mapping = {}
    src_cols = [str(c).strip() for c in df.columns]

    for std_name, keywords in COLUMN_MAPPING.items():
        for kw in keywords:
            for src_col in src_cols:
                if kw.lower() in src_col.lower():
                    mapping[std_name] = src_col
                    break
            if std_name in mapping:
                break

    return mapping


# 业务 ID 列（18 位数字 ID 超过 float64 精度 2^53，必须保持字符串）
ID_COLS = ("Plan ID", "Unit ID", "Message ID")


def _normalize_id_columns(df: pd.DataFrame) -> pd.DataFrame:
    """ID 列去不可见字符（上游 Message ID 末尾常带 \\t）"""
    for col in ID_COLS:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    return df


def _coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """将数值列统一转为 float64（跳过业务 ID 列）"""
    for col in df.columns:
        if col in ID_COLS:
            # 18 位 ID 不能转 float64（精度爆炸），保持字符串
            continue
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].astype("float64")
        else:
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().sum() > len(df) * 0.5:
                nan_count = converted.isna().sum() - df[col].isna().sum()
                if nan_count > 0:
                    logger.warning("列 '%s' 有 %d 个非数值被转为 NaN", col, nan_count)
                df[col] = converted.astype("float64")
    return df


def _apply_column_mapping(df: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    """根据映射重命名列，并确保标准列存在"""
    rename_dict = {v: k for k, v in col_map.items()}
    df = df.rename(columns=rename_dict)

    # 确保所有标准列存在（缺失的填 None）
    for std_name in COLUMN_MAPPING:
        if std_name not in df.columns:
            df[std_name] = None

    return df


def _parse_dates(df: pd.DataFrame) -> pd.DataFrame:
    """解析日期列"""
    if "发送日期" in df.columns:
        df["发送日期"] = pd.to_datetime(df["发送日期"], errors="coerce")
    return df


def _derive_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """计算衍生指标：CTR、GC转化率、触达率（向量化：分母>0 才算，否则 0）"""
    reach = df["触达成功"]
    clicks = df["点击人次"]
    df["CTR"] = np.where(reach > 0, clicks / reach * 100, 0.0)
    df["GC转化率"] = np.where(clicks > 0, df["订单GC"] / clicks * 100, 0.0)
    est = df["预计触达"]
    df["触达率"] = np.where(est > 0, reach / est * 100, 0.0)
    return df


def add_rate_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """聚合后必须先求和再算率（CTR / GC 转化率）：避免按行先算率再平均的精度坑。
    与 _derive_metrics 的区别：本函数假设 df 是已聚合后的明细（行=聚合键），用 np.where
    保证分母>0 才算率，否则 0.0。"""
    if "触达成功" in df.columns and "点击人次" in df.columns:
        df["CTR"] = np.where(
            df["触达成功"] > 0,
            df["点击人次"] / df["触达成功"] * 100,
            0.0,
        )
    if "订单GC" in df.columns and "点击人次" in df.columns:
        df["GC转化率"] = np.where(
            df["点击人次"] > 0,
            df["订单GC"] / df["点击人次"] * 100,
            0.0,
        )
    return df


def data_is_v2(df: pd.DataFrame) -> bool:
    """判断 df 是否是新版（含 Message ID 的 17 列格式）。
    旧文件 read_data 会把缺失列填 None，所以"列存在 + 有非 None 值"才是真新数据。
    三个 caller（tab_plan._aggregate_plans / tab_bu._aggregate_bu_plans / page.py AI handler）
    都用这同一个判定。"""
    return "Message ID" in df.columns and bool(df["Message ID"].notna().any())


def _normalize_unit_column(df: pd.DataFrame) -> None:
    """预归一化 Unit ID：把 "[NULL]" / "" 视为缺失，便于 groupby 用 nunique 一次性向量化。
    就地修改 df；调用方如不想污染原 df 请先 copy。"""
    if "Unit ID" not in df.columns:
        return
    u = df["Unit ID"].astype(str).str.strip()
    df["_unit_norm"] = u.mask(u.isin(("[NULL]", "")))


def read_data(uploaded_file, file_bytes: bytes = None) -> pd.DataFrame:
    """
    读取上传的 CSV 或 XLSX 文件，返回标准化的 DataFrame。
    file_bytes 可选：传入已读取的字节，避免重复读取流。
    """
    filename = uploaded_file.name.lower()
    if file_bytes is not None:
        uploaded_file = BytesIO(file_bytes)

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        df = _read_xlsx(uploaded_file)
    elif filename.endswith(".csv"):
        df = _read_csv(uploaded_file)
    else:
        raise ValueError(f"不支持的文件格式: {filename}，请上传 .csv 或 .xlsx 文件")

    # Fuzzy column mapping
    col_map = _fuzzy_match_columns(df)
    logger.info("列名映射: %s", col_map)

    # 检查必要字段
    required = ["发送日期", "触达成功", "点击人次"]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(f"缺少必要字段: {missing}，请检查文件列名")

    # 应用映射
    df = _apply_column_mapping(df, col_map)

    # 数值列转换
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 日期解析
    df = _parse_dates(df)

    # 衍生指标
    df = _derive_metrics(df)

    return df


def _open_xlsx_sheets(file_bytes: bytes):
    """一次性打开 XLSX workbook，调用方负责 wb.close()。"""
    import openpyxl
    return openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)


def _read_xlsx(uploaded_file) -> pd.DataFrame:
    """读取 XLSX 文件第一个 sheet"""
    wb = _open_xlsx_sheets(uploaded_file.read())
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("XLSX 文件没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    df = _normalize_id_columns(df)
    df = _coerce_numeric_columns(df)

    return df


def read_dau_sheet(file_bytes: bytes) -> pd.DataFrame:
    """
    读取 XLSX 的第二个 sheet（按天去重 DAU）。

    支持两种格式（自动识别）：
      - 旧格式（2 列）：日期 / DAU
      - 新格式（3 列）：日期 / 渠道 / DAU；仅保留 渠道=ALL/all 的行作为总 DAU

    返回列固定为 [日期, DAU]，下游无需感知列数差异。
    接受 bytes（由调用方从 uploaded.read() 获得），避免重复读取流。
    """
    wb = _open_xlsx_sheets(file_bytes)
    if len(wb.sheetnames) < 2:
        wb.close()
        return pd.DataFrame()

    ws = wb[wb.sheetnames[1]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    # 自动检测：3 列格式需 r[1] 为非空字符串（渠道名），而非数字
    has_channel = any(
        len(r) >= 3 and r[1] is not None and isinstance(r[1], str)
        for r in rows[1:]
    )

    if has_channel:
        # 3 列格式：仅保留 渠道=ALL/all 行，输出 [日期, DAU]
        data_rows = [
            (r[0], r[2])
            for r in rows[1:]
            if len(r) >= 3
            and r[0] is not None and r[1] is not None and r[2] is not None
            and str(r[1]).strip().upper() == "ALL"
        ]
    else:
        # 旧格式：2 列直接取
        data_rows = [
            (r[0], r[1])
            for r in rows[1:]
            if len(r) >= 2 and r[0] is not None and r[1] is not None
        ]

    if not data_rows:
        return pd.DataFrame()

    dates, daus = zip(*data_rows)
    df = pd.DataFrame({
        "日期": pd.to_datetime(list(dates), errors="coerce"),
        "DAU": pd.to_numeric(list(daus), errors="coerce"),
    })

    # 去掉空行
    df = df.dropna(subset=["日期", "DAU"])
    df = df.sort_values("日期").reset_index(drop=True)

    return df


def _read_csv(uploaded_file) -> pd.DataFrame:
    """读取 CSV 文件，多编码尝试"""
    bytes_data = uploaded_file.read()

    for enc in ENCODINGS:
        try:
            df = pd.read_csv(BytesIO(bytes_data), encoding=enc, on_bad_lines="skip")
            _normalize_id_columns(df)
            return df
        except Exception:
            continue

    raise ValueError("无法读取 CSV 文件，请检查文件格式和编码")


def filter_week_data(df: pd.DataFrame, week_start=None, week_end=None) -> pd.DataFrame:
    """
    筛选指定周的数据。
    如果未指定日期范围，自动取最近一个自然周（周一~周日）。
    """
    if "发送日期" not in df.columns or df["发送日期"].isna().all():
        raise ValueError("没有有效的发送日期数据")

    df = df.dropna(subset=["发送日期"])

    if week_start is None or week_end is None:
        # 自动取最近一个自然周
        latest_date = df["发送日期"].max()
        # 找到最近的周一
        days_since_monday = latest_date.weekday()
        week_end = latest_date.normalize()
        week_start = week_end - pd.Timedelta(days=days_since_monday)

    week_start = pd.Timestamp(week_start)
    week_end = pd.Timestamp(week_end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    mask = (df["发送日期"] >= week_start) & (df["发送日期"] <= week_end)
    return df.loc[mask].copy()
